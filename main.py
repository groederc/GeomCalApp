import io
from fpdf import FPDF
import geolib as geo
import matplotlib.pyplot as plt
import numpy as npy
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st


# ---------------------------------------------------------------------------
# GENERACIÓN DE EXCEL Y PDF (EN MEMORIA CON FPDF2)
# ---------------------------------------------------------------------------
def generar_excel(datos):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propiedades"

    header_fill = PatternFill("solid", fgColor="1F497D")
    section_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=11, bold=True, color="1F497D")

    ws["A1"] = "REPORTE DE PROPIEDADES GEOMÉTRICAS - FIC-UCOL"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F497D")

    ws.append([])
    ws.append(["Propiedad", "Símbolo", "Valor"])

    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font

    filas = [
        ("Ejes Originales", "", ""),
        ("Área", "A", datos["Area"]),
        ("Centroide X", "Xg", datos["Xg"]),
        ("Centroide Y", "Yg", datos["Yg"]),
        ("Inercia X", "Ix", datos["Ix"]),
        ("Inercia Y", "Iy", datos["Iy"]),
        ("Producto Inercia XY", "Ixy", datos["Ixy"]),
        ("Ejes Centroidales", "", ""),
        ("Inercia Centroidal X", "Ixg", datos["Ixg"]),
        ("Inercia Centroidal Y", "Iyg", datos["Iyg"]),
        ("Producto Inercia Centroidal", "Ixyg", datos["Ixyg"]),
        ("Ejes Principales", "", ""),
        ("Inercia Principal Máx", "I1", datos["Imax"]),
        ("Inercia Principal Mín", "I2", datos["Imin"]),
        ("Ángulo Principal", "φ (°)", datos["phi"]),
    ]

    for item in filas:
        ws.append(list(item))
        r = ws.max_row
        if item[1] == "":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(row=r, column=1).fill = section_fill
            ws.cell(row=r, column=1).font = section_font
        else:
            cell_val = ws.cell(row=r, column=3)
            cell_val.number_format = (
                "0.000E+00"
                if abs(cell_val.value) > 1000 or (0 < abs(cell_val.value) < 0.01)
                else "0.000"
            )

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            max_len + 4, 12
        )

    wb.save(output)
    return output.getvalue()


class PDFReport(FPDF):

    def header(self):
        self.set_fill_color(30, 58, 138)
        self.rect(0, 0, 210, 25, "F")
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(255, 255, 255)
        self.set_xy(10, 6)
        self.cell(0, 7, "Reporte de Propiedades Geometricas - FIC-UCOL", ln=1)
        self.set_font("Helvetica", "", 9)
        self.cell(0, 4, "Calculo de Secciones y Momentos de Inercia", ln=1)
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Pagina {self.page_no()}", align="C")


def generar_pdf(datos):
    pdf = PDFReport()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    secciones = [
        (
            "Ejes Originales",
            [
                ("Area", "A", f"{datos['Area']:.3e}"),
                ("Centroide X", "Xg", f"{datos['Xg']:.3e}"),
                ("Centroide Y", "Yg", f"{datos['Yg']:.3e}"),
                ("Inercia X", "Ix", f"{datos['Ix']:.3e}"),
                ("Inercia Y", "Iy", f"{datos['Iy']:.3e}"),
                ("Producto Inercia XY", "Ixy", f"{datos['Ixy']:.3e}"),
            ],
        ),
        (
            "Ejes Centroidales",
            [
                ("Inercia Centroidal X", "Ixg", f"{datos['Ixg']:.3e}"),
                ("Inercia Centroidal Y", "Iyg", f"{datos['Iyg']:.3e}"),
                ("Producto Inercia Centroidal", "Ixyg", f"{datos['Ixyg']:.3e}"),
            ],
        ),
        (
            "Ejes Principales",
            [
                ("Inercia Principal Maxima", "I1", f"{datos['Imax']:.3e}"),
                ("Inercia Principal Minima", "I2", f"{datos['Imin']:.3e}"),
                ("Angulo Principal", "phi", f"{datos['phi']:.2f} deg"),
            ],
        ),
    ]

    for titulo, filas in secciones:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(30, 58, 138)
        pdf.cell(0, 7, titulo, ln=1)

        pdf.set_fill_color(241, 245, 249)
        pdf.set_text_color(51, 65, 85)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(80, 6, "Propiedad", border=1, fill=True)
        pdf.cell(40, 6, "Simbolo", border=1, fill=True, align="C")
        pdf.cell(60, 6, "Valor", border=1, fill=True, align="R", ln=1)

        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(30, 41, 59)
        for prop, sim, val in filas:
            pdf.cell(80, 6, prop, border=1)
            pdf.cell(40, 6, sim, border=1, align="C")
            pdf.cell(60, 6, val, border=1, align="R", ln=1)
        pdf.ln(4)

    out = pdf.output()
    return bytes(out) if isinstance(out, (bytes, bytearray)) else out.encode("latin1")


# ---------------------------------------------------------------------------
# LÓGICA Y CÁLCULOS
# ---------------------------------------------------------------------------
def ejecutar(edited_df):
    n = edited_df.shape[0]
    st.markdown(f"**Número de filas en la tabla:** {n}")

    if n < 2:
        st.warning(
            "Se requieren al menos 2 puntos para calcular las propiedades."
        )
        return

    try:
        x = edited_df.iloc[:, 0].to_numpy(dtype=float)
        y = edited_df.iloc[:, 1].to_numpy(dtype=float)
    except IndexError:
        st.error(
            "La tabla debe contener al menos 2 columnas numéricas (X e Y)."
        )
        return

    # Cálculos geométricos
    Area, Xg, Yg, Ix, Iy, Ixy, Ixg, Iyg, Ixyg = (
        geo.calculaPropiedadesCentroidales(x, y)
    )
    phi, Imax, Imin = geo.calculaInerciasPrincipales(Ixg, Iyg, Ixyg)

    # Gráfica
    fig, ax = plt.subplots(figsize=(6, 5))
    ax.plot(
        x, y, marker="o", linestyle="-", color="#1f77b4", label="Sección Plana"
    )
    ax.plot(
        Xg,
        Yg,
        marker="P",
        color="red",
        markersize=9,
        label=f"Centroide ({Xg:.2f}, {Yg:.2f})",
    )

    L = max(npy.ptp(x), npy.ptp(y), 1.0) * 0.4
    phi_rad = npy.radians(phi)

    dx1, dy1 = L * npy.cos(phi_rad), L * npy.sin(phi_rad)
    ax.plot(
        [Xg - dx1, Xg + dx1],
        [Yg - dy1, Yg + dy1],
        color="#d62728",
        linestyle="--",
        linewidth=1.5,
        label=f"Eje Principal 1 (φ={phi:.1f}°)",
    )

    dx2, dy2 = -L * npy.sin(phi_rad), L * npy.cos(phi_rad)
    ax.plot(
        [Xg - dx2, Xg + dx2],
        [Yg - dy2, Yg + dy2],
        color="#2ca02c",
        linestyle="--",
        linewidth=1.5,
        label="Eje Principal 2",
    )

    ax.set_aspect("equal")
    ax.set_title(
        "Geometría de la Sección y Ejes Principales",
        fontsize=11,
        fontweight="bold",
    )
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.grid(True, linestyle="--", alpha=0.5)
    ax.legend(loc="upper right", fontsize=8)

    st.pyplot(fig)
    plt.close(fig)

    # Resultados
    st.subheader("Resultados")

    st.markdown("#### 📍 Propiedades respecto a ejes originales")
    c1, c2, c3 = st.columns(3)
    c1.metric("Área", f"{Area:.3e}")
    c2.metric("Centroide X (Xg)", f"{Xg:.3e}")
    c3.metric("Centroide Y (Yg)", f"{Yg:.3e}")

    c4, c5, c6 = st.columns(3)
    c4.metric("Inercia Ix", f"{Ix:.3e}")
    c5.metric("Inercia Iy", f"{Iy:.3e}")
    c6.metric("Producto Inercia Ixy", f"{Ixy:.3e}")

    st.divider()

    st.markdown("#### 🎯 Propiedades respecto a ejes centroidales")
    cg1, cg2, cg3 = st.columns(3)
    cg1.metric("Inercia Ixg", f"{Ixg:.3e}")
    cg2.metric("Inercia Iyg", f"{Iyg:.3e}")
    cg3.metric("Producto Inercia Ixyg", f"{Ixyg:.3e}")

    st.divider()

    st.markdown("#### 📐 Propiedades principales")
    cp1, cp2, cp3 = st.columns(3)
    cp1.metric("Inercia Máx (I1)", f"{Imax:.3e}")
    cp2.metric("Inercia Mín (I2)", f"{Imin:.3e}")
    cp3.metric("Ángulo Principal (φ)", f"{phi:.2f}°")

    # Exportables
    datos_calculados = {
        "Area": Area,
        "Xg": Xg,
        "Yg": Yg,
        "Ix": Ix,
        "Iy": Iy,
        "Ixy": Ixy,
        "Ixg": Ixg,
        "Iyg": Iyg,
        "Ixyg": Ixyg,
        "Imax": Imax,
        "Imin": Imin,
        "phi": phi,
    }

    excel_bytes = generar_excel(datos_calculados)
    pdf_bytes = generar_pdf(datos_calculados)

    st.divider()
    st.subheader("📥 Exportar Resultados")
    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
        st.download_button(
            label="📊 Descargar en Excel (.xlsx)",
            data=excel_bytes,
            file_name="Reporte_Propiedades_Geometricas.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            use_container_width=True,
        )

    with col_dl2:
        st.download_button(
            label="📄 Descargar en PDF (.pdf)",
            data=pdf_bytes,
            file_name="Reporte_Propiedades_Geometricas.pdf",
            mime="application/pdf",
            use_container_width=True,
        )


# ---------------------------------------------------------------------------
# INTERFAZ PRINCIPAL
# ---------------------------------------------------------------------------
def main():
    st.header(
        "Cálculo de propiedades geométricas de figuras planas, FIC-UCOL (2026)"
    )
    st.subheader("Tabla de coordenadas de nudos")
    st.text(
        "Introducir las coordenadas de los vértices de la figura en sentido"
        " antihorario.\nPara cerrar la figura, incluir al final el primer"
        " vértice.\nEl programa automáticamente calcula los valores al ir"
        " introduciendo datos.\n\nAutor:\n    Dr. Guillermo M. Roeder Carbo"
    )

    st.markdown("### Tabla de coordenadas de vértices de la figura")
    df0 = pd.DataFrame([{"X": 0.0, "Y": 0.0}])
    edited_df = st.data_editor(df0, num_rows="dynamic")

    if st.button("Ejecutar", type="primary"):
        ejecutar(edited_df)


if __name__ == "__main__":
    main()
